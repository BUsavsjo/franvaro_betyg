import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import re
import os

# === SÖKVÄGAR ===
DATA_MAPP = os.path.join(os.path.dirname(__file__), "..", "data")
FIL_IN = os.path.join(DATA_MAPP, "franvaro.xls")
FIL_UT = os.path.join(DATA_MAPP, "franvaro_rensad_kategoriserad.xlsx")

# === STEG 1: Läs in och tolka ===
raw = pd.read_excel(FIL_IN, header=None)
data = []
aktuell_klass = None

for row in raw.itertuples(index=False):
    cell = str(row[1]) if len(row) > 1 else ""
    if "Klass:" in cell:
        aktuell_klass = cell.split(":")[1].strip()
    elif pd.to_numeric(row[2], errors="coerce") is not None and aktuell_klass:
        data.append([aktuell_klass] + list(row[1:]))

# Skapa DataFrame
df = pd.DataFrame(data).iloc[:, :11]
df.columns = [
    "klass", "namn", "personnr", "undv_tid", "lekt",
    "n_min", "gf_min", "f_min", "n_pct", "gf_pct", "f_pct"
]

# Ta bort rader där personnr är tom eller det står "namn", "personnr", "undv_tid"
df = df[
    df["personnr"].notna() &
    ~df["personnr"].astype(str).str.lower().str.contains("personnr|namn|undv_tid")
]

# Lägg till årskurs
def extrahera_årskurs(klass):
    if isinstance(klass, str):
        klass = klass.strip()
        if klass.lower().startswith("agsä"):
            return klass
        match = re.search(r"\d", klass)
        if match:
            return f"Åk {match.group()}"
    return "Åk F"

df["årskurs"] = df["klass"].apply(extrahera_årskurs)

# Konvertera procent
def convert_percent(col):
    return pd.to_numeric(
        col.astype(str)
        .str.replace("%", "", regex=False)
        .str.replace(",", ".", regex=False)
        .str.replace("\xa0", "", regex=False)
        .str.extract(r"(\d+\.?\d*)")[0],
        errors='coerce'
    )

df["närvaro_pct"] = convert_percent(df["n_pct"])
df["ogiltig_frånvaro_pct"] = convert_percent(df["f_pct"])

# Ta bort helt tomma mätvärden
df = df[~(
    df["närvaro_pct"].isna() &
    df["ogiltig_frånvaro_pct"].isna()
)]

# === STEG 2: Summering per årskurs ===
def get_total_kategori(narvaro_pct):
    frånvaro = 100 - narvaro_pct if pd.notna(narvaro_pct) else None
    if frånvaro is None:
        return None
    if frånvaro > 50.0: return "50,1--%"
    elif frånvaro > 30.0: return "30,1-50,0%"
    elif frånvaro > 15.0: return "15,1-30,0%"
    elif frånvaro > 5.0: return "5,1-15,0%"
    else: return "0,0-5,0%"

def get_ogiltig_kategori(p):
    if p > 15.0: return "15,1--%"
    elif p > 5.0: return "5,1-15,0%"
    elif p >= 1.0: return "1,0-5,0%"
    else: return None

årskurser = sorted(df["årskurs"].unique())
summering = pd.DataFrame(index=årskurser)

totalkategorier = ["0,0-5,0%", "5,1-15,0%", "15,1-30,0%", "30,1-50,0%", "50,1--%"]
ogiltigkategorier = ["1,0-5,0%", "5,1-15,0%", "15,1--%"]

for kat in totalkategorier:
    summering[f"Total frånvaro {kat}"] = 0
for kat in ogiltigkategorier:
    summering[f"Ogiltig frånvaro {kat}"] = 0
summering["Elevantal"] = 0

for _, row in df.iterrows():
    ak = row["årskurs"]
    total_kat = get_total_kategori(row["närvaro_pct"])
    ogiltig_kat = get_ogiltig_kategori(row["ogiltig_frånvaro_pct"])
    if total_kat:
        summering.at[ak, f"Total frånvaro {total_kat}"] += 1
    if ogiltig_kat:
        summering.at[ak, f"Ogiltig frånvaro {ogiltig_kat}"] += 1
    summering.at[ak, "Elevantal"] += 1

# === STEG 3: Skapa Excel ===
wb = Workbook()
ws1 = wb.active
ws1.title = "Rensad data"

for r in dataframe_to_rows(df, index=False, header=True):
    if any(str(cell).strip() not in ["", "nan", "NaN"] for cell in r):
        ws1.append(r)

# Formatering flik 1
fill_colors = {'A': "FFFFFF", 'B': "C0C0C0", 'C': "C4D79B", 'D': "FFFF99", 'E': "FF9999"}
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=min(5, ws1.max_column)):
    for cell in row:
        col_letter = cell.column_letter
        cell.fill = PatternFill(start_color=fill_colors.get(col_letter, "FFFFFF"),
                                end_color=fill_colors.get(col_letter, "FFFFFF"),
                                fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=(cell.col_idx == 1))
        cell.border = border

# Justera kolumnbredd
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = max_len + 2

# Flik 2: Summering
ws2 = wb.create_sheet("Översikt per årskurs")
for r in dataframe_to_rows(summering, index=True, header=True):
    ws2.append(r)

# Spara
wb.save(FIL_UT)
print(f"✔️ Klar! Filen '{FIL_UT}' har sparats.")
