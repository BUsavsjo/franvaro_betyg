# busavsjo_korrelation_betyg_franvaro.py (uppdaterad version)
import pandas as pd
import hashlib
import json
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from config_paths import OUTPUT_DIR, LASAR

# === INSTÄLLNINGAR ===
DATA_MAPP = OUTPUT_DIR
FIL_FRANVARO = DATA_MAPP / "franvaro_rensad_kategoriserad.xlsx"
FRANVARO_FLIK = "Rensad data"
BETYG_FIL = DATA_MAPP / "betyg.xlsx"
OUTPUT_FIL = DATA_MAPP / "busavsjo_korrelation_input.xlsx"
RESULTAT_FIL = DATA_MAPP / "busavsjo_korrelation_resultat.xlsx"
JSON_MAPP = DATA_MAPP / "json"

IGNORERA_KOLUMNER = {
    "System", "Datum", "Version", "Skolenhetskod",
    "Klass", "Förnamn", "Efternamn", "PersonNr",
}
BETYGSKOLUMNER = None
MODERNA_SPRÅK_KOLUMNER = {"M1(betyg)", "M2(betyg)"}

# === FUNKTIONER ===
def normalisera_personnummer(pnr):
    try:
        return str(pnr).strip()
    except:
        return None

def skapa_anonym_id(pnr):
    return hashlib.sha256(str(pnr).encode('utf-8')).hexdigest()

def mappa_betyg_till_siffra(df, kolumner):
    betygsskala = ['F', 'E', 'D', 'C', 'B', 'A']
    betyg_poang = {b: i for i, b in enumerate(betygsskala)}
    for kol in kolumner:
        ny_kol = kol + "_num"
        df[ny_kol] = df[kol].map(betyg_poang)
    return df

def tolka_korrelation(k):
    if pd.isna(k):
        return "okänd"
    k_abs = abs(k)
    if k_abs >= 0.7:
        styrka = "Tydlig koppling"
    elif k_abs >= 0.4:
        styrka = "Måttlig koppling"
    elif k_abs >= 0.2:
        styrka = "Svag koppling"
    else:
        styrka = "Obetydlig eller ingen koppling"
    riktning = "Negativ (mer frånvaro → lägre betyg)" if k < 0 else "Positiv (mer frånvaro → högre betyg)"
    return f"{styrka} – {riktning}"

def farg_gradient(k):
    if pd.isna(k):
        return None
    k = max(min(k, 1), -1)
    if k < 0:
        r = int(244 + (255 - 244) * (k + 1))
        g = int(204 + (255 - 204) * (k + 1))
        b = int(204 + (255 - 204) * (k + 1))
    else:
        r = int(255 - (255 - 217) * k)
        g = int(255 - (255 - 234) * k)
        b = int(255 - (255 - 211) * k)
    return PatternFill(start_color=f"{r:02X}{g:02X}{b:02X}", fill_type="solid")

# === STEG 1 ===
betyg_df = pd.read_excel(BETYG_FIL, dtype={"PersonNr": str})
BETYGSKOLUMNER = [kol for kol in betyg_df.columns if kol not in IGNORERA_KOLUMNER]

betyg_df["AnonymID"] = (
    betyg_df["PersonNr"].apply(normalisera_personnummer).apply(skapa_anonym_id)
)
betyg_df.drop(columns=[c for c in IGNORERA_KOLUMNER if c in betyg_df.columns], inplace=True)
betyg_df = mappa_betyg_till_siffra(betyg_df, BETYGSKOLUMNER)

# === STEG 2 ===
franvaro_df = pd.read_excel(FIL_FRANVARO, sheet_name=FRANVARO_FLIK)
franvaro_df = franvaro_df[franvaro_df["personnr"].notna()]
franvaro_df["AnonymID"] = franvaro_df["personnr"].apply(normalisera_personnummer).apply(skapa_anonym_id)
franvaro_df = franvaro_df[["AnonymID", "årskurs", "ogiltig_frånvaro_pct", "närvaro_pct"]]
franvaro_df["total_frånvaro"] = 100 - franvaro_df["närvaro_pct"]

# === STEG 3 ===
samman_df = pd.merge(betyg_df, franvaro_df, on="AnonymID")

# === Infoga meritvärde före export ===
betygspoang = {'F': 0, 'E': 10, 'D': 12.5, 'C': 15, 'B': 17.5, 'A': 20}

def meritvarde(rad):
    poang = [betygspoang.get(b, 0) for b in rad if pd.notna(b)]
    topp = 17 if MODERNA_SPRÅK_KOLUMNER.intersection(BETYGSKOLUMNER) else 16
    return sum(sorted(poang, reverse=True)[:topp])

samman_df["meritvarde"] = samman_df[BETYGSKOLUMNER].apply(meritvarde, axis=1)

# === STEG 4 ===
korrelationer = []
for betyg in [f"{amne}_num" for amne in BETYGSKOLUMNER]:
    for kol in ["ogiltig_frånvaro_pct", "total_frånvaro"]:
        if betyg in samman_df.columns and kol in samman_df.columns:
            x, y = samman_df[betyg], samman_df[kol]
            if x.notna().sum() >= 2 and y.notna().sum() >= 2:
                corr = x.corr(y)
                if pd.notna(corr):
                    korrelationer.append({
                        "Ämne": betyg.replace("_num", ""),
                        "Frånvarotyp": kol,
                        "Korrelation": round(corr, 2),
                        "Styrka": tolka_korrelation(corr)
                    })

# === STEG 5 ===
samman_df.to_excel(OUTPUT_FIL, index=False)
# === STEG 6 ===
if korrelationer:
    resultat_df = pd.DataFrame(korrelationer)
    resultat_df["Läsår"] = LASAR
    ogiltig_df = resultat_df[resultat_df["Frånvarotyp"] == "ogiltig_frånvaro_pct"]
    total_df = resultat_df[resultat_df["Frånvarotyp"] == "total_frånvaro"]

    with pd.ExcelWriter(RESULTAT_FIL, engine='openpyxl') as writer:
        ogiltig_df.to_excel(writer, index=False, sheet_name="Ogiltig frånvaro")
        total_df.to_excel(writer, index=False, sheet_name="Total frånvaro")

    wb = load_workbook(RESULTAT_FIL)
    for bladnamn in ["Ogiltig frånvaro", "Total frånvaro"]:
        ws = wb[bladnamn]
        for rad in range(2, ws.max_row + 1):
            cell = ws.cell(row=rad, column=3)
            try:
                value = float(cell.value)
                fill = farg_gradient(value)
                if fill:
                    cell.fill = fill
            except:
                continue
    wb.save(RESULTAT_FIL)

# === STEG 7 ===
JSON_MAPP.mkdir(parents=True, exist_ok=True)

def spara_json(df, filnamn):
    df_clean = df.copy()
    df_clean["Korrelation"] = df_clean["Korrelation"].apply(lambda x: round(x, 2) if pd.notna(x) else None)
    df_clean["Läsår"] = LASAR
    df_clean = df_clean.where(pd.notna(df_clean), None)
    with (JSON_MAPP / filnamn).open("w", encoding="utf-8") as f:
        json.dump(df_clean.to_dict(orient="records"), f, indent=2, ensure_ascii=False)

spara_json(ogiltig_df, "ogiltig_franvaro.json")
spara_json(total_df, "total_franvaro.json")

# === STEG 8 ===
print("\n📊 Korrelation mellan MERITVÄRDE och frånvaro:\n")
betygspoang = {'F': 0, 'E': 10, 'D': 12.5, 'C': 15, 'B': 17.5, 'A': 20}

def meritvarde(rad):
    poang = [betygspoang.get(b, 0) for b in rad if pd.notna(b)]
    topp = 17 if MODERNA_SPRÅK_KOLUMNER.intersection(BETYGSKOLUMNER) else 16
    return sum(sorted(poang, reverse=True)[:topp])

samman_df["meritvarde"] = samman_df[BETYGSKOLUMNER].apply(meritvarde, axis=1)

korrelation_merit = []
for kol in ["ogiltig_frånvaro_pct", "total_frånvaro"]:
    x, y = samman_df["meritvarde"], samman_df[kol]
    if x.notna().sum() >= 2 and y.notna().sum() >= 2:
        k = x.corr(y)
        korrelation_merit.append({
            "Typ": kol,
            "Korrelation": round(k, 2),
            "Styrka": tolka_korrelation(k)
        })

# === STEG 9 ===
if korrelation_merit:
    merit_df = pd.DataFrame(korrelation_merit)
    merit_df["Läsår"] = LASAR
    with pd.ExcelWriter(RESULTAT_FIL, engine='openpyxl', mode='a') as writer:
        merit_df.to_excel(writer, index=False, sheet_name="Meritvärde vs frånvaro")

    wb = load_workbook(RESULTAT_FIL)
    ws = wb["Meritvärde vs frånvaro"]
    for rad in range(2, ws.max_row + 1):
        cell = ws.cell(row=rad, column=2)
        try:
            value = float(cell.value)
            fill = farg_gradient(value)
            if fill:
                cell.fill = fill
        except:
            continue
    wb.save(RESULTAT_FIL)

    for rad in korrelation_merit:
        rad["Läsår"] = LASAR
        filnamn = "merit_" + ("ogiltig" if "ogiltig" in rad["Typ"] else "total") + "_franvaro.json"
        with (JSON_MAPP / filnamn).open("w", encoding="utf-8") as f:
            json.dump(rad, f, indent=2, ensure_ascii=False)
